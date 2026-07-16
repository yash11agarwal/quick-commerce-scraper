"""Excel workbook (job_tracker.xlsx) as the tool's entire user interface.

One workbook is both the config INPUT and the pipeline OUTPUT:

    input sheets   Searches / Filters / Scoring / Settings — yours to edit,
                   never rewritten by the tool
    output sheets  Jobs + Stats — regenerated from SQLite on every
                   hunt/sync; don't reorder their columns

The Jobs sheet is also (partially) an input: the ``Status`` column has a
dropdown and the ``Add Note`` column is free text. On the next
``jobs.py hunt`` / ``jobs.py sync`` those edits are applied to the store
(status changes recorded in history, notes appended with a timestamp) and
the sheet is rebuilt — Add Note cells come back empty once absorbed.

SQLite stays the source of truth; the workbook is a view plus an inbox
of your edits. Save & close the file in Excel before running a sync.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .config import (EXPERIENCE_LEVELS, JOB_TYPES, POSTED_WITHIN,
                     WORKPLACE_TYPES, FilterConfig, JobHunterConfig,
                     RateLimitConfig, RetryConfig, SearchSpec,
                     validate_searches)
from .schema import JobStatus, utcnow_iso
from .store import JobStore

log = logging.getLogger(__name__)

DEFAULT_WORKBOOK = "job_tracker.xlsx"

SHEET_README = "ReadMe"
SHEET_SEARCHES = "Searches"
SHEET_FILTERS = "Filters"
SHEET_SCORING = "Scoring"
SHEET_SETTINGS = "Settings"
SHEET_JOBS = "Jobs"
SHEET_STATS = "Stats"

SEARCH_HEADERS = ["Name", "Keywords", "Location", "Posted Within",
                  "Experience Levels", "Workplace", "Job Types", "Max Pages"]
FILTER_HEADERS = ["Title Must Contain (any)", "Title Exclude",
                  "Company Exclude"]
SCORING_HEADERS = ["Keyword", "Points"]
SETTINGS_HEADERS = ["Setting", "Value", "What it does"]
JOBS_HEADERS = ["Job ID", "Title", "Company", "Location", "Posted", "Search",
                "Score", "Status", "Add Note", "Notes Log", "URL",
                "First Seen", "Last Seen", "Status Updated"]

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", start_color="DDEBF7")
_NEW_ROW_FILL = PatternFill("solid", start_color="FFF2CC")  # light yellow

_STATUS_ORDER = {s: i for i, s in enumerate(JobStatus.values())}


# --------------------------------------------------------------------------
# template creation
# --------------------------------------------------------------------------

_README_LINES = [
    "LinkedIn job hunter & tracker — this workbook is the whole interface.",
    "",
    "Input sheets (yours; the tool only reads them):",
    "  Searches  one row per saved LinkedIn search",
    "  Filters   drop rules applied before a job is stored",
    "  Scoring   keyword -> points; ranks jobs (top score first)",
    "  Settings  politeness / storage knobs",
    "",
    "Output sheets (rebuilt from the database on every hunt/sync):",
    "  Jobs      your pipeline. TWO columns are editable:",
    "            - Status: pick from the dropdown (applied, interviewing...)",
    "            - Add Note: type anything; it's saved as a timestamped",
    "              note on the next run, then the cell is cleared",
    "  Stats     funnel counts",
    "",
    "Commands (run from the repo root):",
    "  python jobs.py hunt     pull new postings + absorb your Jobs edits",
    "  python jobs.py sync     absorb edits / refresh sheets, no hunting",
    "  python jobs.py show <id> --fetch    full job description in terminal",
    "",
    "Allowed values (comma-separate multiples inside one cell):",
    f"  Posted Within:      {', '.join(POSTED_WITHIN)}",
    f"  Experience Levels:  {', '.join(EXPERIENCE_LEVELS)}",
    f"  Workplace:          {', '.join(WORKPLACE_TYPES)}",
    f"  Job Types:          {', '.join(JOB_TYPES)}",
    f"  Status:             {', '.join(JobStatus.values())}",
    "",
    "Save and CLOSE the workbook before running a hunt/sync — the tool",
    "can't write while Excel holds the file open.",
]

_SAMPLE_SEARCHES = [
    ["python-bangalore", "python developer", "Bengaluru, Karnataka, India",
     "week", "entry, associate", "", "", 3],
    ["data-remote", "data engineer", "India",
     "week", "", "remote", "", 3],
]

_SAMPLE_FILTERS = [
    ["", "principal", ""],
    ["", "staff ", ""],
    ["", "architect", ""],
]

_SAMPLE_SCORING = [
    ["python", 3], ["backend", 2], ["django", 2], ["fastapi", 2],
    ["scraping", 2], ["remote", 1],
]

_SETTINGS_ROWS = [
    ["min_delay_seconds", 5, "minimum gap between requests to LinkedIn"],
    ["jitter_seconds", 3, "random 0..N seconds added on top of the gap"],
    ["max_attempts", 3, "retries per request before giving up"],
    ["backoff_base_seconds", 5, "retry backoff: 5s, 10s, 20s..."],
    ["request_timeout_seconds", 30, "per-request HTTP timeout"],
    ["sqlite_path", "data/jobs.db", "where the tracking database lives"],
]


def _style_header(ws: Worksheet, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


def _set_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def _dropdown(ws: Worksheet, values: Iterable[str], cell_range: str) -> None:
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"',
                        allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def create_template(path: str | Path) -> None:
    """Write a fresh job_tracker.xlsx with sample inputs and empty outputs."""
    wb = Workbook()

    readme = wb.active
    readme.title = SHEET_README
    for i, line in enumerate(_README_LINES, start=1):
        readme.cell(row=i, column=1, value=line)
    _set_widths(readme, {"A": 100})

    ws = wb.create_sheet(SHEET_SEARCHES)
    ws.append(SEARCH_HEADERS)
    for row in _SAMPLE_SEARCHES:
        ws.append(row)
    _style_header(ws, len(SEARCH_HEADERS))
    _set_widths(ws, {"A": 18, "B": 28, "C": 30, "D": 14, "E": 22, "F": 16,
                     "G": 16, "H": 10})
    _dropdown(ws, POSTED_WITHIN, "D2:D200")
    ws["E1"].comment = Comment(
        "Comma-separated: " + ", ".join(EXPERIENCE_LEVELS), "job_hunter")
    ws["F1"].comment = Comment(
        "Comma-separated: " + ", ".join(WORKPLACE_TYPES), "job_hunter")
    ws["G1"].comment = Comment(
        "Comma-separated: " + ", ".join(JOB_TYPES), "job_hunter")

    ws = wb.create_sheet(SHEET_FILTERS)
    ws.append(FILTER_HEADERS)
    for row in _SAMPLE_FILTERS:
        ws.append(row)
    _style_header(ws, len(FILTER_HEADERS))
    _set_widths(ws, {"A": 26, "B": 26, "C": 26})

    ws = wb.create_sheet(SHEET_SCORING)
    ws.append(SCORING_HEADERS)
    for row in _SAMPLE_SCORING:
        ws.append(row)
    _style_header(ws, len(SCORING_HEADERS))
    _set_widths(ws, {"A": 22, "B": 8})

    ws = wb.create_sheet(SHEET_SETTINGS)
    ws.append(SETTINGS_HEADERS)
    for row in _SETTINGS_ROWS:
        ws.append(row)
    _style_header(ws, len(SETTINGS_HEADERS))
    _set_widths(ws, {"A": 26, "B": 14, "C": 50})

    _init_jobs_sheet(wb.create_sheet(SHEET_JOBS))
    _init_stats_sheet(wb.create_sheet(SHEET_STATS))

    wb.save(path)


# --------------------------------------------------------------------------
# config input: workbook -> JobHunterConfig
# --------------------------------------------------------------------------

def _clean(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _split_multi(value) -> list[str]:
    text = _clean(value)
    if not text:
        return []
    return [part.strip().lower() for part in text.split(",") if part.strip()]


def _rows_as_dicts(ws: Worksheet) -> list[dict]:
    headers = [_clean(c.value) for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {h: v for h, v in zip(headers, row) if h}
        if any(v is not None and str(v).strip() for v in rec.values()):
            out.append(rec)
    return out


def _require_sheet(wb, name: str, path) -> Worksheet:
    if name not in wb.sheetnames:
        raise ValueError(
            f"{path}: sheet {name!r} is missing — regenerate the workbook "
            "with `python jobs.py init --force` (back up your data first) "
            "or restore the sheet")
    return wb[name]


def _column_values(ws: Worksheet, header: str) -> list[str]:
    headers = [_clean(c.value) for c in ws[1]]
    if header not in headers:
        return []
    idx = headers.index(header)
    values = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = _clean(row[idx]) if idx < len(row) else None
        if v:
            values.append(v)
    return values


def load_config(path: str | Path) -> JobHunterConfig:
    """Read the input sheets of job_tracker.xlsx into a JobHunterConfig."""
    wb = load_workbook(path, data_only=True)

    searches = []
    for rec in _rows_as_dicts(_require_sheet(wb, SHEET_SEARCHES, path)):
        name = _clean(rec.get("Name"))
        keywords = _clean(rec.get("Keywords"))
        location = _clean(rec.get("Location"))
        if not (name and keywords and location):
            raise ValueError(
                f"{path}: every Searches row needs Name, Keywords and "
                f"Location (offending row: {rec})")
        posted = _clean(rec.get("Posted Within"))
        searches.append(SearchSpec(
            name=name, keywords=keywords, location=location,
            posted_within=posted.lower() if posted else None,
            experience_levels=_split_multi(rec.get("Experience Levels")),
            workplace=_split_multi(rec.get("Workplace")),
            job_types=_split_multi(rec.get("Job Types")),
            max_pages=int(rec.get("Max Pages") or 4),
        ))
    validate_searches(searches, str(path))

    fws = _require_sheet(wb, SHEET_FILTERS, path)
    filters = FilterConfig(
        title_include_any=_column_values(fws, FILTER_HEADERS[0]),
        title_exclude=_column_values(fws, FILTER_HEADERS[1]),
        company_exclude=_column_values(fws, FILTER_HEADERS[2]),
    )

    score_keywords: dict[str, int] = {}
    for rec in _rows_as_dicts(_require_sheet(wb, SHEET_SCORING, path)):
        kw = _clean(rec.get("Keyword"))
        if kw:
            score_keywords[kw.lower()] = int(rec.get("Points") or 1)

    settings = {r["Setting"]: r.get("Value")
                for r in _rows_as_dicts(_require_sheet(wb, SHEET_SETTINGS, path))
                if _clean(r.get("Setting"))}

    def _num(key: str, default: float) -> float:
        value = settings.get(key)
        return float(value) if value is not None else default

    return JobHunterConfig(
        searches=searches,
        filters=filters,
        score_keywords=score_keywords,
        rate_limit=RateLimitConfig(
            min_delay_seconds=_num("min_delay_seconds", 5.0),
            jitter_seconds=_num("jitter_seconds", 3.0),
        ),
        retry=RetryConfig(
            max_attempts=int(_num("max_attempts", 3)),
            backoff_base_seconds=_num("backoff_base_seconds", 5.0),
        ),
        sqlite_path=_clean(settings.get("sqlite_path")) or "data/jobs.db",
        request_timeout_seconds=_num("request_timeout_seconds", 30.0),
    )


# --------------------------------------------------------------------------
# Jobs-sheet edits: workbook -> store
# --------------------------------------------------------------------------

@dataclass
class JobEdit:
    job_id: str
    status: Optional[str]
    add_note: Optional[str]


def read_job_edits(path: str | Path) -> list[JobEdit]:
    """Collect (status, add-note) values per row of the Jobs sheet."""
    wb = load_workbook(path, data_only=True)
    if SHEET_JOBS not in wb.sheetnames:
        return []
    edits = []
    for rec in _rows_as_dicts(wb[SHEET_JOBS]):
        job_id = _clean(rec.get("Job ID"))
        if not job_id:
            continue
        status = _clean(rec.get("Status"))
        edits.append(JobEdit(
            job_id=job_id,
            status=status.lower() if status else None,
            add_note=_clean(rec.get("Add Note")),
        ))
    return edits


def apply_edits(store: JobStore, edits: list[JobEdit]) -> tuple[int, int, list[str]]:
    """Apply Jobs-sheet edits to the store.

    Returns (status_changes, notes_added, warnings). Bad rows warn and are
    skipped — a typo in Excel must never abort a hunt.
    """
    status_changes = notes_added = 0
    warnings: list[str] = []
    for edit in edits:
        try:
            job = store.get_job(edit.job_id)
        except KeyError:
            warnings.append(f"Jobs sheet row with unknown id {edit.job_id!r} "
                            "— ignored (row was edited or DB was reset?)")
            continue
        if edit.status and edit.status != job["status"]:
            try:
                store.set_status(job["job_id"], edit.status)
                status_changes += 1
            except ValueError:
                warnings.append(
                    f"{job['job_id']}: invalid status {edit.status!r} in Jobs "
                    f"sheet (valid: {', '.join(JobStatus.values())}) — kept "
                    f"{job['status']!r}")
        if edit.add_note:
            store.add_note(job["job_id"], edit.add_note)
            notes_added += 1
    return status_changes, notes_added, warnings


# --------------------------------------------------------------------------
# pipeline output: store -> workbook
# --------------------------------------------------------------------------

def _init_jobs_sheet(ws: Worksheet) -> None:
    ws.append(JOBS_HEADERS)
    _style_header(ws, len(JOBS_HEADERS))
    _set_widths(ws, {"A": 12, "B": 40, "C": 24, "D": 22, "E": 11, "F": 16,
                     "G": 7, "H": 13, "I": 28, "J": 50, "K": 42, "L": 20,
                     "M": 20, "N": 20})
    _dropdown(ws, JobStatus.values(), "H2:H5000")
    ws["H1"].comment = Comment(
        "Editable — pick a status; it's saved on the next hunt/sync.",
        "job_hunter")
    ws["I1"].comment = Comment(
        "Editable — anything typed here is stored as a timestamped note on "
        "the next hunt/sync, then cleared.", "job_hunter")


def _write_jobs_ws(ws: Worksheet, rows, new_ids: set[str]) -> None:
    _init_jobs_sheet(ws)
    ordered = sorted(rows, key=lambda r: (_STATUS_ORDER.get(r["status"], 99),
                                          -r["score"]))
    for r in ordered:
        ws.append([
            r["job_id"], r["title"], r["company"], r["location"],
            r["posted_date"], r["search_name"], r["score"], r["status"],
            None,  # Add Note inbox, cleared after every sync
            r["notes"].strip() if r["notes"] else None,
            r["url"], r["first_seen"], r["last_seen"], r["status_updated_at"],
        ])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=1).number_format = "@"
        url_cell = ws.cell(row=row_idx, column=11)
        url_cell.hyperlink = r["url"]
        url_cell.style = "Hyperlink"
        notes_cell = ws.cell(row=row_idx, column=10)
        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if r["job_id"] in new_ids:
            for col in range(1, len(JOBS_HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = _NEW_ROW_FILL
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:N{ws.max_row}"


def _init_stats_sheet(ws: Worksheet) -> None:
    ws.append(["Pipeline", "Count"])
    _style_header(ws, 2)
    _set_widths(ws, {"A": 22, "B": 10})


def _write_stats_ws(ws: Worksheet, store: JobStore) -> None:
    _init_stats_sheet(ws)
    counts = store.stats()
    for status, n in counts.items():
        ws.append([status, n])
    ws.append(["total", sum(counts.values())])
    ws.cell(row=ws.max_row, column=1).font = _HEADER_FONT
    ws.append([])
    header_row = ws.max_row + 1
    ws.append(["By search", "Count"])
    for col in (1, 2):
        cell = ws.cell(row=header_row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for name, n in store.stats_by_search().items():
        ws.append([name, n])
    ws.append([])
    ws.append([f"updated {utcnow_iso()}"])


def _replace_sheet(wb, name: str) -> Worksheet:
    if name in wb.sheetnames:
        index = wb.sheetnames.index(name)
        del wb[name]
        return wb.create_sheet(name, index)
    return wb.create_sheet(name)


def write_tracker_sheets(path: str | Path, store: JobStore,
                         new_ids: Iterable[str] = ()) -> None:
    """Rebuild the Jobs and Stats sheets from the store (input sheets and
    everything else in the workbook are left untouched)."""
    wb = load_workbook(path)
    _write_jobs_ws(_replace_sheet(wb, SHEET_JOBS), store.list_jobs(statuses=[]),
                   set(new_ids))
    _write_stats_ws(_replace_sheet(wb, SHEET_STATS), store)
    wb.save(path)


def export_snapshot(path: str | Path, store: JobStore) -> int:
    """Standalone .xlsx dump of the pipeline (Jobs + Stats). Returns rows."""
    wb = Workbook()
    jobs_ws = wb.active
    jobs_ws.title = SHEET_JOBS
    rows = store.list_jobs(statuses=[])
    _write_jobs_ws(jobs_ws, rows, set())
    _write_stats_ws(wb.create_sheet(SHEET_STATS), store)
    wb.save(path)
    return len(rows)
