#!/usr/bin/env python3
"""Generate targets.xlsx — the fill-in-the-blanks workbook for exact-SKU tracking.

Run this only if you need to regenerate the template (e.g. it got deleted or
you want a fresh copy). It overwrites targets.xlsx in the project root.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def main() -> None:
    wb = Workbook()

    # --- Instructions sheet (first tab, so it's what opens by default) ---
    instr = wb.active
    instr.title = "Instructions"
    instr.column_dimensions["A"].width = 100
    lines = [
        "HOW TO USE THIS FILE",
        "",
        "This lets the scraper track the EXACT product you mean, instead of "
        "guessing from a text search (which can grab the wrong pack size or brand).",
        "",
        "STEP 1 — Fill in the 'Pincodes' tab",
        "List every delivery pincode you want checked, one per row.",
        "",
        "STEP 2 — Fill in the 'Products' tab",
        "For every product you want tracked, on every platform that sells it:",
        "  1. Open that platform's website or app in your normal browser.",
        "  2. Search for the product and CLICK INTO IT (open its own product page).",
        "  3. Copy the full web address from your browser's address bar.",
        "  4. Paste it into the 'product_url' column.",
        "  5. In 'platform', type exactly one of: blinkit, swiggy_instamart, zepto, "
        "bigbasket_now, amazon_now, flipkart_minutes",
        "  6. In 'label', type a short name YOU choose, e.g. 'Amul Butter 500g'. "
        "Use the SAME label for the same product across different platforms — "
        "that's what lets the dashboard compare their prices side by side.",
        "",
        "STEP 3 — Run it",
        "In PowerShell, in your project folder, run:",
        "    python track_products.py --headed -v",
        "",
        "A visible browser will open and fetch each exact product page you listed. "
        "Results are saved to the same database the regular scraper uses, so "
        "'python dashboard.py' shows this data too.",
        "",
        "If a row fails, check the ./debug/<platform>/ folder that gets created — "
        "it saves a screenshot of what went wrong.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = instr.cell(row=i, column=1, value=line)
        cell.alignment = WRAP
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif line.startswith(("STEP",)):
            cell.font = Font(bold=True, size=12)

    # --- Pincodes sheet ---
    pin = wb.create_sheet("Pincodes")
    pin.append(["pincode"])
    style_header(pin, 1)
    pin.column_dimensions["A"].width = 20
    for p in ("110001", "560001"):
        pin.append([p])

    # --- Products sheet ---
    prod = wb.create_sheet("Products")
    prod.append(["platform", "product_url", "label"])
    style_header(prod, 3)
    prod.column_dimensions["A"].width = 20
    prod.column_dimensions["B"].width = 70
    prod.column_dimensions["C"].width = 25
    sample_rows = [
        ("blinkit", "https://blinkit.com/prn/amul-butter-500-g/prid/123456",
         "Amul Butter 500g"),
        ("bigbasket_now", "https://www.bigbasket.com/pd/40023/amul-butter-500-g/",
         "Amul Butter 500g"),
        ("", "", ""),
        ("", "", "<-- delete the sample rows above and add your own"),
    ]
    for row in sample_rows:
        prod.append(row)

    out = Path(__file__).resolve().parent.parent / "targets.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
