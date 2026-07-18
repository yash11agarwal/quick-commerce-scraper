#!/usr/bin/env python3
"""Generate inputs.xlsx — the fill-in-the-blanks workbook for keyword search.

Sheet "Products" = product search terms (one per row); sheet "Pincodes" =
delivery pincodes (one per row). When inputs.xlsx exists in the project
root, main.py uses it automatically instead of the lists in config.yaml.

Run this only to regenerate the template (it overwrites inputs.xlsx).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def main() -> None:
    wb = Workbook()

    instr = wb.active
    instr.title = "Instructions"
    instr.column_dimensions["A"].width = 100
    lines = [
        "HOW TO USE THIS FILE",
        "",
        "This file tells the scraper WHAT to search for and WHERE. When this "
        "file exists next to main.py, it is used automatically — you don't "
        "need to touch config.yaml for products/pincodes anymore.",
        "",
        "STEP 1 — 'Products' tab: type one product search term per row "
        "(e.g. 'amul butter 500g'), just like you'd type it into the app's search box.",
        "",
        "STEP 2 — 'Pincodes' tab: type one delivery pincode per row.",
        "",
        "STEP 3 — Save this file, then run in PowerShell:",
        "    python main.py --headed -v",
        "",
        "Every enabled platform will be searched for every product at every "
        "pincode. Results go into the database; view them with "
        "'python dashboard.py' — which also has a 'Download Excel' button — "
        "or export directly with 'python export_excel.py'.",
        "",
        "TIP: for tracking an EXACT product (not a search), use targets.xlsx "
        "with 'python track_products.py' instead — see its Instructions tab.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = instr.cell(row=i, column=1, value=line)
        cell.alignment = WRAP
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif line.startswith("STEP"):
            cell.font = Font(bold=True, size=12)

    prod = wb.create_sheet("Products")
    prod.append(["product"])
    style_header(prod, 1)
    prod.column_dimensions["A"].width = 40
    for sample in ("amul butter 500g", "maggi noodles", "coca cola 750ml"):
        prod.append([sample])

    pin = wb.create_sheet("Pincodes")
    pin.append(["pincode"])
    style_header(pin, 1)
    pin.column_dimensions["A"].width = 20
    for p in ("110001", "560001"):
        pin.append([p])

    out = Path(__file__).resolve().parent.parent / "inputs.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
