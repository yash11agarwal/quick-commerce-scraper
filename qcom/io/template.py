"""Write a blank input workbook with the expected sheets and headers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

_INSTRUCTIONS = [
    "How to fill this workbook",
    "",
    "Sheet 'products': one search term per row in product_name, exactly as you would type it on the app.",
    "  brand and pack_size are optional and only improve match_score; category is passed through to the output.",
    "  active: leave blank or TRUE to include the row, FALSE to park it.",
    "",
    "Sheet 'pincodes': one six-digit pincode per row. The column is formatted as text; keep it that way.",
    "  city and state are optional. They make the location check stricter. active works as above.",
    "",
    "Sheet 'settings' (optional): platforms as a comma-separated list (blank = every implemented platform),",
    "  max_results_per_query as a whole number (default 20), run_label as free text written into the output.",
    "",
    "Then run:  python -m qcom run --input input.xlsx --out output/",
]


def write_template(path: str | Path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(["product_name", "brand", "pack_size", "category", "active"])
    ws2 = wb.create_sheet("pincodes")
    ws2.append(["pincode", "city", "state", "active"])
    ws2.column_dimensions["A"].number_format = "@"  # whole column as text so typed pincodes keep leading zeros
    ws3 = wb.create_sheet("settings")
    ws3.append(["key", "value"])
    ws3.append(["platforms", ""])
    ws3.append(["max_results_per_query", 20])
    ws3.append(["run_label", ""])
    ws4 = wb.create_sheet("README")
    for line in _INSTRUCTIONS:
        ws4.append([line])
    ws4["A1"].font = Font(bold=True)
    ws4.column_dimensions["A"].width = 110
    for sheet in (ws, ws2, ws3):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["B"].width = 18
    wb.save(p)
    return p
