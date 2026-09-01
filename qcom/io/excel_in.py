"""Read and validate the input workbook. Every problem is reported with sheet and cell.

Sheet 1 holds products, sheet 2 holds pincodes (found by name first, then by position). A
``settings`` sheet is optional. Nothing is coerced silently: a bad pincode is a validation
error naming the cell, never a guess.
"""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from qcom.core.errors import InputProblem, InputValidationError
from qcom.core.models import PINCODE_RE, InputSpec, PincodeInput, ProductInput, RunSettings

PRODUCT_ALIASES: dict[str, tuple[str, ...]] = {
    "product_name": ("product_name", "product name", "product", "name", "search_term", "search term", "query"),
    "brand": ("brand",),
    "pack_size": ("pack_size", "pack size", "pack", "size"),
    "category": ("category",),
    "active": ("active",),
}
PINCODE_ALIASES: dict[str, tuple[str, ...]] = {
    "pincode": ("pincode", "pin_code", "pin code", "pin", "postal code", "postcode"),
    "city": ("city",),
    "state": ("state",),
    "active": ("active",),
}
SETTINGS_KEYS = ("platforms", "max_results_per_query", "run_label")
_NON_DATA_SHEET_PREFIXES = ("settings", "readme", "instruction", "help", "notes")


def _text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _header_map(ws: Worksheet, aliases: dict[str, tuple[str, ...]], sheet_label: str, problems: list[InputProblem]) -> dict[str, int]:
    """Column index (1-based) per canonical field, from row 1."""
    found: dict[str, int] = {}
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for idx, cell in enumerate(row1, start=1):
        label = _text(cell).lower()
        if not label:
            continue
        for canonical, names in aliases.items():
            if label in names:
                if canonical in found:
                    problems.append(InputProblem(sheet_label, f"{get_column_letter(idx)}1", f"duplicate column for {canonical!r}"))
                found[canonical] = idx
    return found


def _active(value: Any, sheet: str, cell: str, problems: list[InputProblem]) -> bool:
    text = _text(value).upper()
    if text in ("", "TRUE", "YES", "Y", "1"):
        return True
    if text in ("FALSE", "NO", "N", "0"):
        return False
    problems.append(InputProblem(sheet, cell, f"active must be TRUE or FALSE, got {value!r}"))
    return False


def _pincode_text(value: Any, sheet: str, cell: str, problems: list[InputProblem]) -> str | None:
    if value is None or _text(value) == "":
        problems.append(InputProblem(sheet, cell, "pincode is blank"))
        return None
    if isinstance(value, bool):
        problems.append(InputProblem(sheet, cell, f"pincode must be six digits, got {value!r}"))
        return None
    if isinstance(value, float):
        if not value.is_integer():
            problems.append(InputProblem(sheet, cell, f"pincode must be six digits, got {value!r}"))
            return None
        value = int(value)
    text = str(value).strip()
    if not PINCODE_RE.match(text):
        problems.append(InputProblem(sheet, cell, f"pincode must be exactly six digits as text, got {text!r}"))
        return None
    return text


def _pick_sheets(wb: openpyxl.Workbook) -> tuple[Worksheet | None, Worksheet | None, Worksheet | None]:
    products = pincodes = settings = None
    for name in wb.sheetnames:
        low = name.strip().lower()
        if products is None and low.startswith("product"):
            products = wb[name]
        elif pincodes is None and (low.startswith("pincode") or low.startswith("pin")):
            pincodes = wb[name]
        elif settings is None and low.startswith("settings"):
            settings = wb[name]
    positional = [wb[n] for n in wb.sheetnames if not n.strip().lower().startswith(_NON_DATA_SHEET_PREFIXES)]
    positional = [ws for ws in positional if ws is not products and ws is not pincodes]
    if products is None and positional:
        products = positional.pop(0)
    if pincodes is None and positional:
        pincodes = positional.pop(0)
    return products, pincodes, settings


def read_input(path: str | Path) -> InputSpec:
    p = Path(path)
    if not p.exists():
        raise InputValidationError([InputProblem("(file)", "-", f"{p} does not exist")])
    try:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001 - reported as a validation problem with the cause
        raise InputValidationError([InputProblem("(file)", "-", f"cannot open {p}: {exc}")]) from exc

    problems: list[InputProblem] = []
    ws_products, ws_pincodes, ws_settings = _pick_sheets(wb)
    if ws_products is None:
        problems.append(InputProblem("(workbook)", "-", "no products sheet: sheet 1 must hold product names"))
    if ws_pincodes is None:
        problems.append(InputProblem("(workbook)", "-", "no pincodes sheet: sheet 2 must hold pincodes"))
    if problems:
        raise InputValidationError(problems)
    assert ws_products is not None and ws_pincodes is not None

    # ---- products
    ps = ws_products.title
    pcols = _header_map(ws_products, PRODUCT_ALIASES, ps, problems)
    if "product_name" not in pcols:
        problems.append(InputProblem(ps, "A1", "missing header: expected a 'product_name' column (aliases: product, name, search_term)"))
    products: list[ProductInput] = []
    seen_products: dict[str, int] = {}
    if "product_name" in pcols:
        for row_idx, row in enumerate(ws_products.iter_rows(min_row=2, values_only=True), start=2):
            if all(_text(c) == "" for c in row):
                continue

            def col(field: str) -> Any:
                i = pcols.get(field)
                return row[i - 1] if i is not None and i - 1 < len(row) else None

            name = _text(col("product_name"))
            cell = f"{get_column_letter(pcols['product_name'])}{row_idx}"
            active = _active(col("active"), ps, f"{get_column_letter(pcols['active'])}{row_idx}", problems) if "active" in pcols else True
            if not active:
                continue
            if not name:
                problems.append(InputProblem(ps, cell, "product_name is blank on an active row"))
                continue
            key = " ".join(name.lower().split())
            if key in seen_products:
                problems.append(InputProblem(ps, cell, f"duplicate product_name {name!r} (first seen in row {seen_products[key]})"))
                continue
            seen_products[key] = row_idx
            products.append(
                ProductInput(
                    input_row_id=row_idx,
                    product_name=name,
                    brand=_text(col("brand")) or None,
                    pack_size=_text(col("pack_size")) or None,
                    category=_text(col("category")) or None,
                )
            )

    # ---- pincodes
    zs = ws_pincodes.title
    zcols = _header_map(ws_pincodes, PINCODE_ALIASES, zs, problems)
    if "pincode" not in zcols:
        problems.append(InputProblem(zs, "A1", "missing header: expected a 'pincode' column"))
    pincodes: list[PincodeInput] = []
    seen_pins: dict[str, int] = {}
    if "pincode" in zcols:
        for row_idx, row in enumerate(ws_pincodes.iter_rows(min_row=2, values_only=True), start=2):
            if all(_text(c) == "" for c in row):
                continue

            def zcol(field: str) -> Any:
                i = zcols.get(field)
                return row[i - 1] if i is not None and i - 1 < len(row) else None

            cell = f"{get_column_letter(zcols['pincode'])}{row_idx}"
            active = _active(zcol("active"), zs, f"{get_column_letter(zcols['active'])}{row_idx}", problems) if "active" in zcols else True
            if not active:
                continue
            pin = _pincode_text(zcol("pincode"), zs, cell, problems)
            if pin is None:
                continue
            if pin in seen_pins:
                problems.append(InputProblem(zs, cell, f"duplicate pincode {pin} (first seen in row {seen_pins[pin]})"))
                continue
            seen_pins[pin] = row_idx
            pincodes.append(PincodeInput(input_row_id=row_idx, pincode=pin, city=_text(zcol("city")) or None, state=_text(zcol("state")) or None))

    # ---- settings
    raw_settings: dict[str, str] = {}
    if ws_settings is not None:
        ss = ws_settings.title
        for row_idx, row in enumerate(ws_settings.iter_rows(min_row=1, values_only=True), start=1):
            if not row or all(_text(c) == "" for c in row):
                continue
            key = _text(row[0]).lower()
            value = _text(row[1]) if len(row) > 1 else ""
            if key in ("key", "setting"):
                continue
            if key not in SETTINGS_KEYS:
                problems.append(InputProblem(ss, f"A{row_idx}", f"unknown setting {key!r}; known: {', '.join(SETTINGS_KEYS)}"))
                continue
            raw_settings[key] = value
    platforms = [s.strip() for s in raw_settings.get("platforms", "").split(",") if s.strip()]
    max_results = 20
    if raw_settings.get("max_results_per_query"):
        try:
            max_results = int(raw_settings["max_results_per_query"])
            if max_results < 1:
                raise ValueError
        except ValueError:
            problems.append(InputProblem(ws_settings.title if ws_settings else "settings", "-", f"max_results_per_query must be a positive integer, got {raw_settings['max_results_per_query']!r}"))
    if not problems and not products:
        problems.append(InputProblem(ps, "-", "no active products"))
    if not problems and not pincodes:
        problems.append(InputProblem(zs, "-", "no active pincodes"))
    if problems:
        raise InputValidationError(problems)

    return InputSpec(
        source_path=str(p),
        sha256=hashlib.sha256(p.read_bytes()).hexdigest(),
        products=products,
        pincodes=pincodes,
        settings=RunSettings(platforms=platforms, max_results_per_query=max_results, run_label=raw_settings.get("run_label") or None),
    )
