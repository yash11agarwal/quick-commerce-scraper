"""Write the output workbook from the database. The workbook is a projection, never a source."""

from __future__ import annotations

import json
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from qcom.core.clock import parse_iso, to_ist
from qcom.core.storage import Storage

_HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TEXT = "@"
_MONEY = "0.00"
_PCT = "0.00"
_DATETIME = "yyyy-mm-dd hh:mm:ss"

# (header, kind) kind decides the cell format: text | money | int | pct | datetime | plain
RESULT_COLUMNS: list[tuple[str, str]] = [
    ("run_id", "text"), ("captured_at_ist", "datetime"), ("captured_at_utc", "text"), ("platform", "text"),
    ("requested_pincode", "text"), ("effective_pincode", "text"), ("city", "text"), ("search_term", "text"),
    ("input_row_id", "int"), ("result_rank", "int"), ("platform_product_id", "text"), ("product_name", "text"),
    ("brand", "text"), ("pack_size", "text"), ("unit_normalised", "text"), ("mrp", "money"), ("selling_price", "money"),
    ("base_selling_price", "money"), ("discount_pct", "pct"), ("price_per_unit", "money"), ("currency", "text"),
    ("in_stock", "plain"), ("stock_qty", "int"), ("eta_minutes", "int"), ("store_or_seller_id", "text"),
    ("category_path", "text"), ("product_url", "text"), ("image_url", "text"), ("match_score", "pct"),
    ("raw_payload_ref", "text"), ("strategy", "text"),
]

SUMMARY_COLUMNS: list[tuple[str, str]] = [
    ("platform", "text"), ("requested_pincode", "text"), ("effective_pincode", "text"), ("search_term", "text"),
    ("input_row_id", "int"), ("status", "text"), ("results_returned", "int"), ("attempts", "int"), ("duration_ms", "int"),
    ("strategy", "text"), ("final_code", "text"), ("final_reason", "text"), ("store_id", "text"), ("eta_minutes", "int"),
    ("first_started_utc", "text"), ("last_finished_utc", "text"),
]

FAILURE_COLUMNS: list[tuple[str, str]] = [
    ("job_id", "text"), ("platform", "text"), ("requested_pincode", "text"), ("search_term", "text"), ("input_row_id", "int"),
    ("status", "text"), ("error_code", "text"), ("reason", "text"), ("attempts", "int"), ("last_attempt_utc", "text"),
    ("raw_payload_or_screenshot", "text"),
]


def _paise(v: int | None) -> Decimal | None:
    return None if v is None else (Decimal(v) / Decimal(100)).quantize(Decimal("0.01"))


def _dec(v: str | None) -> Decimal | None:
    return None if v in (None, "") else Decimal(str(v))


def _ist_naive(iso_utc: str | None) -> datetime | None:
    if not iso_utc:
        return None
    return to_ist(parse_iso(iso_utc)).replace(tzinfo=None)


def _write_sheet(ws: Worksheet, columns: list[tuple[str, str]], rows: list[list[Any]]) -> None:
    ws.append([h for h, _ in columns])
    for i in range(1, len(columns) + 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font, c.alignment = _HEADER_FILL, _HEADER_FONT, Alignment(vertical="center")
    for row in rows:
        ws.append(row)
    for col_idx, (header, kind) in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        width = 14
        if header in ("product_name", "final_reason", "reason", "product_url", "image_url", "raw_payload_or_screenshot", "category_path"):
            width = 44
        elif header in ("search_term", "job_id", "captured_at_ist", "captured_at_utc", "store_or_seller_id", "platform_product_id"):
            width = 24
        ws.column_dimensions[letter].width = width
        fmt = {"text": _TEXT, "money": _MONEY, "pct": _PCT, "datetime": _DATETIME}.get(kind)
        if fmt:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = fmt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(ws.max_row, 1)}"


def _results(storage: Storage, run_id: str) -> list[list[Any]]:
    out: list[list[Any]] = []
    for r in storage.results_rows(run_id):
        in_stock = None if r["in_stock"] is None else bool(r["in_stock"])
        out.append(
            [
                r["run_id"], _ist_naive(r["captured_at_utc"]), r["captured_at_utc"], r["platform"], r["requested_pincode"],
                r["effective_pincode"], r["job_city"], r["search_term"], r["input_row_id"], r["result_rank"], r["platform_product_id"],
                r["product_name"], r["brand"], r["pack_size"], r["unit_normalised"], _paise(r["mrp_paise"]),
                _paise(r["selling_price_paise"]), _paise(r["base_selling_price_paise"]), _dec(r["discount_pct"]),
                _paise(r["price_per_unit_paise"]), r["currency"], in_stock, r["stock_qty"], r["eta_minutes"], r["store_or_seller_id"],
                r["category_path"], r["product_url"], r["image_url"], _dec(r["match_score"]), r["capture_id"], r["strategy"],
            ]
        )
    return out


def _summary(storage: Storage, run_id: str) -> list[list[Any]]:
    out: list[list[Any]] = []
    for j in storage.job_rows(run_id):
        out.append(
            [
                j["platform"], j["requested_pincode"], j["effective_pincode"], j["search_term"], j["input_row_id"], j["status"],
                j["results_returned"], j["attempts"], j["duration_ms"], j["strategy"], j["final_code"], j["final_reason"],
                j["store_id"], j["eta_minutes"], j["first_started_utc"], j["last_finished_utc"],
            ]
        )
    return out


def _failures(storage: Storage, run_id: str) -> list[list[Any]]:
    out: list[list[Any]] = []
    for j in storage.job_rows(run_id):
        if j["status"] == "OK":
            continue
        out.append(
            [
                j["job_id"], j["platform"], j["requested_pincode"], j["search_term"], j["input_row_id"], j["status"],
                j["final_code"], j["final_reason"], j["attempts"], j["last_finished_utc"], j["artifact_path"],
            ]
        )
    return out


def _meta(storage: Storage, run_id: str) -> list[list[Any]]:
    run = storage.get_run(run_id) or {}
    summary = json.loads(run.get("summary_json") or "{}")
    rows: list[list[Any]] = [
        ["run_id", run_id],
        ["run_label", run.get("run_label")],
        ["started_at_utc", run.get("started_at_utc")],
        ["started_at_ist", run.get("started_at_ist")],
        ["ended_at_utc", run.get("ended_at_utc")],
        ["ended_at_ist", run.get("ended_at_ist")],
        ["code_version", run.get("code_version")],
        ["git_sha", run.get("git_sha")],
        ["config_hash", run.get("config_hash")],
        ["input_path", run.get("input_path")],
        ["input_sha256", run.get("input_sha256")],
        ["proxy", run.get("proxy_label") or "none"],
        ["run_status", run.get("status")],
        ["exit_code", run.get("exit_code")],
    ]
    for name, version in json.loads(run.get("adapter_versions_json") or "{}").items():
        rows.append([f"adapter_version.{name}", version])
    for k, v in (summary.get("status_counts") or {}).items():
        rows.append([f"jobs.{k}", v])
    for k, v in (summary.get("code_counts") or {}).items():
        rows.append([f"code.{k}", v])
    for k, v in (summary.get("dq_counts") or {}).items():
        rows.append([f"data_quality.{k}", v])
    for s in summary.get("platform_states") or []:
        rows.append([f"platform.{s['platform']}", f"{s['status']}" + (f": {s['reason']}" if s.get("reason") else "")])
    for note in summary.get("notes") or []:
        rows.append(["note", note])
    return rows


def write_workbook(storage: Storage, run_id: str, out_path: str | Path) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    _write_sheet(ws, RESULT_COLUMNS, _results(storage, run_id))
    _write_sheet(wb.create_sheet("run_summary"), SUMMARY_COLUMNS, _summary(storage, run_id))
    _write_sheet(wb.create_sheet("failures"), FAILURE_COLUMNS, _failures(storage, run_id))
    meta = wb.create_sheet("run_meta")
    _write_sheet(meta, [("key", "text"), ("value", "plain")], _meta(storage, run_id))
    meta.column_dimensions["A"].width = 30
    meta.column_dimensions["B"].width = 80
    wb.save(out)
    return out
