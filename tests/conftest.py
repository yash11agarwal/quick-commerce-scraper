from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from qcom.core.config import AppConfig, RetryCfg, RetryEntry, StorageCfg, ThrottleCfg
from qcom.core.models import InputSpec, PincodeInput, ProductInput, RunSettings


def fast_config(tmp_path: Path, **overrides) -> AppConfig:
    cfg = AppConfig(
        throttle=ThrottleCfg(min_gap_s=0, jitter_s=0),
        retry=RetryCfg(
            network_timeout=RetryEntry(attempts=3, backoff_base_s=0.001, jitter_s=0),
            rate_limited=RetryEntry(attempts=2, backoff_base_s=0.001, jitter_s=0),
            proxy_error=RetryEntry(attempts=2, backoff_base_s=0.001, jitter_s=0),
            location_not_set=RetryEntry(attempts=2, backoff_base_s=0.001, jitter_s=0),
            unknown=RetryEntry(attempts=1),
        ),
        storage=StorageCfg(path=str(tmp_path / "db.sqlite"), sessions_dir=str(tmp_path / "sessions"), runs_dir=str(tmp_path / "runs")),
    )
    for key, value in overrides.items():
        setattr(cfg, key, value)
    return cfg


@pytest.fixture
def cfg(tmp_path: Path) -> AppConfig:
    return fast_config(tmp_path)


def spec_for(products: list[str], pincodes: list[str], *, platforms: list[str] | None = None, max_results: int = 8, brand: str | None = None, pack: str | None = None) -> InputSpec:
    return InputSpec(
        source_path="memory.xlsx",
        sha256="0" * 64,
        products=[ProductInput(input_row_id=i + 2, product_name=p, brand=brand, pack_size=pack) for i, p in enumerate(products)],
        pincodes=[PincodeInput(input_row_id=i + 2, pincode=z) for i, z in enumerate(pincodes)],
        settings=RunSettings(platforms=platforms or ["fake"], max_results_per_query=max_results),
    )


@pytest.fixture
def make_workbook(tmp_path: Path):
    """Build an input workbook: rows are tuples; headers are given per sheet."""

    def _make(
        products: list[tuple], pincodes: list[tuple], *,
        product_headers=("product_name", "brand", "pack_size", "category", "active"),
        pincode_headers=("pincode", "city", "state", "active"),
        settings: list[tuple] | None = None,
        sheet_names=("products", "pincodes", "settings"),
        name="input.xlsx",
    ) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_names[0]
        ws.append(list(product_headers))
        for r in products:
            ws.append(list(r))
        wp = wb.create_sheet(sheet_names[1])
        wp.append(list(pincode_headers))
        for r in pincodes:
            wp.append(list(r))
        if settings is not None:
            wsx = wb.create_sheet(sheet_names[2])
            wsx.append(["key", "value"])
            for r in settings:
                wsx.append(list(r))
        path = tmp_path / name
        wb.save(path)
        return path

    return _make
