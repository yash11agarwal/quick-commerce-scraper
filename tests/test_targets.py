import pytest
from openpyxl import Workbook

from qc_scraper.targets import extract_product_id, load_targets

# Real-shaped sample URLs (invented ids) matching each adapter's get_inventory() pattern.
_SAMPLE_URLS = {
    "blinkit": ("https://blinkit.com/prn/amul-butter-500-g/prid/123456", "123456"),
    "zepto": ("https://www.zeptonow.com/pn/amul-butter/pvid/abc-123-def", "abc-123-def"),
    "swiggy_instamart": ("https://www.swiggy.com/instamart/item/xyz987", "xyz987"),
    "bigbasket_now": ("https://www.bigbasket.com/pd/40023/amul-butter-500-g/", "40023"),
    "amazon_now": ("https://www.amazon.in/Amul-Butter/dp/B0EXAMPLE1", "B0EXAMPLE1"),
    "flipkart_minutes": ("https://www.flipkart.com/amul-butter/p/itmabc?pid=MLKGHXYZ123", "MLKGHXYZ123"),
}


@pytest.mark.parametrize("platform", list(_SAMPLE_URLS))
def test_extract_product_id_matches_real_url_shapes(platform):
    url, expected_id = _SAMPLE_URLS[platform]
    assert extract_product_id(platform, url) == expected_id


def test_extract_product_id_rejects_non_product_url():
    with pytest.raises(ValueError, match="couldn't find a product id"):
        extract_product_id("blinkit", "https://blinkit.com/s/?q=amul+butter")


def test_extract_product_id_unknown_platform():
    with pytest.raises(ValueError, match="no URL pattern known"):
        extract_product_id("not_a_real_platform", "https://example.com/x/1")


def _make_workbook(path, pincode_rows, product_rows,
                   pincode_header=("pincode",),
                   product_header=("platform", "product_url", "label")):
    wb = Workbook()
    wb.remove(wb.active)
    pin = wb.create_sheet("Pincodes")
    pin.append(list(pincode_header))
    for row in pincode_rows:
        pin.append(row)
    prod = wb.create_sheet("Products")
    prod.append(list(product_header))
    for row in product_rows:
        prod.append(row)
    wb.save(path)


def test_load_targets_happy_path(tmp_path):
    path = tmp_path / "targets.xlsx"
    _make_workbook(
        path,
        pincode_rows=[["110001"], ["560001"]],
        product_rows=[
            ["blinkit", "https://blinkit.com/prn/x/prid/1", "Amul Butter 500g"],
            ["bigbasket_now", "https://www.bigbasket.com/pd/2/x/", "Amul Butter 500g"],
        ],
    )
    pincodes, products = load_targets(path)
    assert pincodes == ["110001", "560001"]
    assert len(products) == 2
    assert products[0].platform == "blinkit"
    assert products[0].label == "Amul Butter 500g"


def test_load_targets_skips_blank_rows_and_missing_url(tmp_path):
    path = tmp_path / "targets.xlsx"
    _make_workbook(
        path,
        pincode_rows=[["110001"], [None], [""]],
        product_rows=[
            [None, None, None],
            ["blinkit", "", "no url, should be skipped"],
            ["blinkit", "https://blinkit.com/prn/x/prid/9", "Real Row"],
        ],
    )
    pincodes, products = load_targets(path)
    assert pincodes == ["110001"]
    assert len(products) == 1
    assert products[0].label == "Real Row"


def test_load_targets_defaults_label_to_url_when_blank(tmp_path):
    path = tmp_path / "targets.xlsx"
    _make_workbook(
        path,
        pincode_rows=[["110001"]],
        product_rows=[["blinkit", "https://blinkit.com/prn/x/prid/9", ""]],
    )
    _, products = load_targets(path)
    assert products[0].label == "https://blinkit.com/prn/x/prid/9"


def test_load_targets_missing_column_raises_clear_error(tmp_path):
    path = tmp_path / "targets.xlsx"
    _make_workbook(
        path,
        pincode_rows=[["110001"]],
        product_rows=[["blinkit", "https://blinkit.com/prn/x/prid/9", "x"]],
        product_header=("platform", "url_typo", "label"),
    )
    with pytest.raises(ValueError, match="missing column"):
        load_targets(path)


def test_shipped_template_loads_without_the_sample_rows(tmp_path):
    """The real targets.xlsx (with its two example rows) must still parse."""
    from pathlib import Path
    template = Path(__file__).resolve().parent.parent / "targets.xlsx"
    if not template.exists():
        pytest.skip("targets.xlsx not generated in this checkout")
    pincodes, products = load_targets(template)
    assert pincodes  # sample pincodes present
    assert all(p.platform and p.url for p in products)  # blank sample row was skipped
