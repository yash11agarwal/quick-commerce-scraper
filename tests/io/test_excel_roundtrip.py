from datetime import datetime
from decimal import Decimal

import openpyxl
from openpyxl import Workbook

from qcom.core.runner import execute_run, plan_run
from qcom.core.storage import Storage
from qcom.io.excel_out import RESULT_COLUMNS, _write_sheet, write_workbook
from tests.conftest import spec_for


def test_money_precision_and_text_pincode_survive(tmp_path):
    wb = Workbook()
    ws = wb.active
    values = [Decimal("74.25"), Decimal("0.01"), Decimal("1299.99"), Decimal("100000.50"), Decimal("34.00")]
    rows = [["700048", "000123", v] for v in values]
    _write_sheet(ws, [("pincode", "text"), ("id", "text"), ("price", "money")], rows)
    path = tmp_path / "rt.xlsx"
    wb.save(path)
    back = openpyxl.load_workbook(path)["A1"].parent if False else openpyxl.load_workbook(path).active
    for i, v in enumerate(values, start=2):
        assert back.cell(row=i, column=1).value == "700048" and back.cell(row=i, column=1).number_format == "@"
        assert back.cell(row=i, column=2).value == "000123"  # leading zeros intact
        cell = back.cell(row=i, column=3)
        assert Decimal(str(cell.value)) == v, (cell.value, v)
        assert cell.number_format == "0.00"


def test_full_workbook_from_a_fake_run(cfg, tmp_path):
    run_id = plan_run(cfg, spec_for(["amul butter", "nothing at all"], ["700048"], max_results=8))
    summary = execute_run(cfg, run_id)
    assert summary.exit_code == 0
    with Storage(cfg.storage.path) as storage:
        path = write_workbook(storage, run_id, tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["results", "run_summary", "failures", "run_meta"]
    ws = wb["results"]
    headers = [c.value for c in ws[1]]
    assert headers == [h for h, _ in RESULT_COLUMNS]
    assert ws.freeze_panes == "A2" and ws.auto_filter.ref.startswith("A1:")
    row = {h: ws.cell(row=2, column=i + 1) for i, h in enumerate(headers)}
    assert row["requested_pincode"].value == "700048" and row["requested_pincode"].number_format == "@"
    assert row["effective_pincode"].value == "700048"
    assert isinstance(row["captured_at_ist"].value, datetime) and row["captured_at_ist"].number_format == "yyyy-mm-dd hh:mm:ss"
    assert row["captured_at_utc"].value.endswith("+00:00")
    assert Decimal(str(row["selling_price"].value)) == Decimal("275") and row["selling_price"].number_format == "0.00"
    assert Decimal(str(row["base_selling_price"].value)) == Decimal("290")
    assert row["platform_product_id"].value == "F001" and row["in_stock"].value is True and row["stock_qty"].value == 5
    assert row["raw_payload_ref"].value.startswith(run_id) and row["strategy"].value == "fixture_search"
    assert row["currency"].value == "INR"
    # the row with no MRP has blank mrp and blank discount, never 0
    mrp_col = headers.index("mrp") + 1
    disc_col = headers.index("discount_pct") + 1
    blank_mrp_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=mrp_col).value is None]
    assert blank_mrp_rows and all(ws.cell(row=r, column=disc_col).value is None for r in blank_mrp_rows)
    # failures sheet lists the NO_RESULTS job with its code; run_meta carries the counts
    failures = [[c.value for c in r] for r in wb["failures"].iter_rows(min_row=2)]
    assert len(failures) == 1 and failures[0][6] == "NO_RESULTS"
    meta = {r[0].value: r[1].value for r in wb["run_meta"].iter_rows(min_row=2)}
    assert meta["run_id"] == run_id and meta["jobs.OK"] == 1 and meta["jobs.NO_RESULTS"] == 1 and meta["proxy"] == "none"
