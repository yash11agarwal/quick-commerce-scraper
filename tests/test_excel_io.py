import openpyxl
from openpyxl import Workbook

from qc_scraper.excel_io import export_observations, load_search_inputs
from qc_scraper.schema import ProductRecord
from qc_scraper.storage import ObservationStore


def test_load_inputs_by_sheet_name_and_header(tmp_path):
    wb = Workbook()
    wb.active.title = "Instructions"
    prod = wb.create_sheet("Products")
    prod.append(["product"])
    prod.append(["amul butter 500g"])
    prod.append([None])                    # blank row skipped
    prod.append(["maggi noodles"])
    prod.append(["amul butter 500g"])      # duplicate dropped
    pin = wb.create_sheet("Pincodes")
    pin.append(["pincode"])
    pin.append([110001])                   # numeric cell -> clean string
    pin.append(["560001"])
    path = tmp_path / "inputs.xlsx"
    wb.save(path)

    queries, pincodes = load_search_inputs(path)
    assert queries == ["amul butter 500g", "maggi noodles"]
    assert pincodes == ["110001", "560001"]


def test_load_inputs_headerless_positional_sheets(tmp_path):
    # Bare workbook: default sheet names, no header row — sheet 1 products,
    # sheet 2 pincodes, exactly as the user described.
    wb = Workbook()
    s1 = wb.active
    s1.append(["coca cola 750ml"])
    s2 = wb.create_sheet("Sheet2")
    s2.append([700048])
    path = tmp_path / "bare.xlsx"
    wb.save(path)

    queries, pincodes = load_search_inputs(path)
    assert queries == ["coca cola 750ml"]
    assert pincodes == ["700048"]


def _rec(price: float, name="Amul Butter 500 g", pincode="110001") -> ProductRecord:
    return ProductRecord(
        platform="blinkit", product_name=name, brand="Amul", price=price,
        mrp=305.0, discount_pct=None, quantity_unit="500 g", in_stock=True,
        stock_estimate=None, stock_granularity="boolean", pincode=pincode,
        product_id="101", search_query="amul butter 500g",
    ).finalize()


def test_export_roundtrip(tmp_path):
    db = tmp_path / "obs.db"
    with ObservationStore(db) as store:
        store.insert_many([_rec(270.0), _rec(275.0), _rec(280.0, pincode="560001")])

    out = tmp_path / "export.xlsx"
    count = export_observations(db, out)
    assert count == 3

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Latest snapshot", "All observations"]
    all_ws = wb["All observations"]
    assert all_ws.max_row == 4  # header + 3 observations
    header = [c.value for c in all_ws[1]]
    assert "Price (₹)" in header and "In stock" in header
    # Latest snapshot: one row per (platform, product, pincode) = 2 rows
    assert wb["Latest snapshot"].max_row == 3
    assert all_ws.cell(row=2, column=header.index("In stock") + 1).value == "Yes"


def test_export_filters(tmp_path):
    db = tmp_path / "obs.db"
    with ObservationStore(db) as store:
        store.insert_many([_rec(270.0), _rec(280.0, pincode="560001")])
    out = tmp_path / "filtered.xlsx"
    count = export_observations(db, out, pincode="110001")
    assert count == 1
