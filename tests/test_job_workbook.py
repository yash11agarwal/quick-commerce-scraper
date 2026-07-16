"""Round-trip tests for the Excel interface (template -> config -> sheets)."""

import pytest
from openpyxl import load_workbook

from job_hunter import workbook
from job_hunter.schema import JobRecord
from job_hunter.store import JobStore


@pytest.fixture
def wb_path(tmp_path):
    path = tmp_path / "job_tracker.xlsx"
    workbook.create_template(path)
    return path


@pytest.fixture
def store(tmp_path):
    with JobStore(tmp_path / "jobs.db") as s:
        yield s


def _job(job_id, title="Python Developer", score=3) -> JobRecord:
    return JobRecord(
        job_id=job_id, title=title, company="Acme", location="Bengaluru",
        url=f"https://www.linkedin.com/jobs/view/{job_id}",
        posted_date="2026-07-14", search_name="py-blr", score=score)


def test_template_loads_as_config(wb_path):
    cfg = workbook.load_config(wb_path)

    assert [s.name for s in cfg.searches] == ["python-bangalore", "data-remote"]
    first = cfg.searches[0]
    params = first.query_params()
    assert params["f_TPR"] == "r604800"          # week
    assert params["f_E"] == "2,3"                # entry, associate
    assert cfg.searches[1].query_params()["f_WT"] == "2"  # remote

    assert "principal" in cfg.filters.title_exclude
    assert cfg.score_keywords["python"] == 3
    assert cfg.sqlite_path == "data/jobs.db"
    assert cfg.rate_limit.min_delay_seconds == 5


def test_config_validation_from_sheet(wb_path):
    wb = load_workbook(wb_path)
    ws = wb[workbook.SHEET_SEARCHES]
    ws["D2"] = "fortnight"                        # invalid posted_within
    wb.save(wb_path)
    with pytest.raises(ValueError, match="posted_within"):
        workbook.load_config(wb_path)


def test_jobs_sheet_roundtrip_absorbs_edits(wb_path, store):
    store.add_jobs([_job("4012345678"), _job("4098765432", "Data Engineer", 1)])
    workbook.write_tracker_sheets(wb_path, store, new_ids={"4012345678"})

    # The user works the pipeline in Excel: status dropdown + a note.
    wb = load_workbook(wb_path)
    ws = wb[workbook.SHEET_JOBS]
    headers = [c.value for c in ws[1]]
    id_col = headers.index("Job ID") + 1
    status_col = headers.index("Status") + 1
    note_col = headers.index("Add Note") + 1
    row = next(r for r in range(2, ws.max_row + 1)
               if str(ws.cell(r, id_col).value) == "4012345678")
    ws.cell(row, status_col, "applied")
    ws.cell(row, note_col, "sent CV via portal")
    wb.save(wb_path)

    edits = workbook.read_job_edits(wb_path)
    changed, noted, warnings = workbook.apply_edits(store, edits)
    assert (changed, noted, warnings) == (1, 1, [])
    job = store.get_job("4012345678")
    assert job["status"] == "applied"
    assert "sent CV via portal" in job["notes"]

    # Rebuild: the absorbed note inbox is cleared, status persists,
    # and Stats reflects the pipeline.
    workbook.write_tracker_sheets(wb_path, store)
    wb = load_workbook(wb_path)
    ws = wb[workbook.SHEET_JOBS]
    values = {str(ws.cell(r, id_col).value):
              (ws.cell(r, status_col).value, ws.cell(r, note_col).value)
              for r in range(2, ws.max_row + 1)}
    assert values["4012345678"] == ("applied", None)
    assert values["4098765432"] == ("new", None)
    stats_rows = [tuple(r) for r in
                  wb[workbook.SHEET_STATS].iter_rows(values_only=True)]
    assert ("new", 1) in stats_rows and ("applied", 1) in stats_rows


def test_apply_edits_survives_bad_rows(wb_path, store):
    store.add_jobs([_job("4012345678")])
    edits = [
        workbook.JobEdit("9999999999", "applied", None),   # unknown id
        workbook.JobEdit("4012345678", "ghosted", "hi"),   # invalid status
    ]
    changed, noted, warnings = workbook.apply_edits(store, edits)
    assert changed == 0
    assert noted == 1                       # the note still lands
    assert len(warnings) == 2
    assert store.get_job("4012345678")["status"] == "new"


def test_input_sheets_survive_refresh(wb_path, store):
    # Customizing an input sheet must never be undone by a sheet rebuild.
    wb = load_workbook(wb_path)
    wb[workbook.SHEET_SCORING]["A8"] = "golang"
    wb[workbook.SHEET_SCORING]["B8"] = 5
    wb.save(wb_path)

    store.add_jobs([_job("4012345678")])
    workbook.write_tracker_sheets(wb_path, store)

    cfg = workbook.load_config(wb_path)
    assert cfg.score_keywords["golang"] == 5


def test_export_snapshot(tmp_path, store):
    store.add_jobs([_job("4012345678"), _job("4098765432")])
    out = tmp_path / "snapshot.xlsx"
    assert workbook.export_snapshot(out, store) == 2
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {workbook.SHEET_JOBS, workbook.SHEET_STATS}
    assert wb[workbook.SHEET_JOBS].max_row == 3  # header + 2 jobs
