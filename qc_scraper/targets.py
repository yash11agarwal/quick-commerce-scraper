"""Load exact-SKU tracking targets from an Excel workbook.

This is the alternative to keyword search in config.yaml: instead of typing
a search term and hoping the scraper's parser picks the right result out of
a results page, you paste the *exact* product page URL you want tracked —
copied straight from your own browser's address bar. See the README's
"Exact-SKU tracking" section for how to fill in targets.xlsx.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass
class TargetProduct:
    platform: str
    url: str
    label: str


# How to pull each platform's internal product id out of a URL a human
# copy-pasted from their own browser. These mirror the URL shape each
# adapter's get_inventory() already builds — see the get_inventory() method
# in the matching qc_scraper/scrapers/<platform>.py for the canonical shape.
# REVERSE-ENGINEERED, same caveat as everywhere else in this project: if a
# platform changes its URL scheme, these patterns need updating too.
_ID_PATTERNS: dict[str, re.Pattern] = {
    "blinkit": re.compile(r"/prid/(\d+)"),
    "zepto": re.compile(r"/pvid/([\w-]+)"),
    "swiggy_instamart": re.compile(r"/instamart/item/([\w-]+)"),
    "bigbasket_now": re.compile(r"/pd/(\w+)"),
    "amazon_now": re.compile(r"/dp/([A-Z0-9]{10})", re.IGNORECASE),
    "flipkart_minutes": re.compile(r"[?&]pid=([A-Za-z0-9]+)"),
}


def extract_product_id(platform: str, url: str) -> str:
    pattern = _ID_PATTERNS.get(platform)
    if pattern is None:
        raise ValueError(
            f"no URL pattern known for platform {platform!r} — check for a typo "
            f"in the Products sheet (valid: {sorted(_ID_PATTERNS)})"
        )
    match = pattern.search(url)
    if not match:
        raise ValueError(
            f"couldn't find a product id in this {platform} URL: {url!r} — make "
            "sure you copied the full product PAGE url (after clicking into the "
            "product), not a search results or category link"
        )
    return match.group(1)


def load_targets(path: str | Path) -> tuple[list[str], list[TargetProduct]]:
    """Read the Pincodes and Products sheets from the targets workbook.

    Both sheets are read by header name (row 1), case-insensitively, so
    column order doesn't matter and blank rows are skipped — the sheet is
    meant to be hand-edited in Excel/Google Sheets by someone non-technical.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    pincodes: list[str] = []
    if "Pincodes" in wb.sheetnames:
        ws = wb["Pincodes"]
        header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        if "pincode" not in header:
            raise ValueError('Pincodes sheet needs a "pincode" column header in row 1')
        col = header.index("pincode")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if col < len(row) and row[col]:
                pincodes.append(str(row[col]).strip())

    products: list[TargetProduct] = []
    if "Products" in wb.sheetnames:
        ws = wb["Products"]
        header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        required = ("platform", "product_url", "label")
        missing = [c for c in required if c not in header]
        if missing:
            raise ValueError(f"Products sheet is missing column(s): {missing}")
        idx = {c: header.index(c) for c in required}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            platform = str(row[idx["platform"]] or "").strip()
            url = str(row[idx["product_url"]] or "").strip()
            label = str(row[idx["label"]] or "").strip()
            if not platform or not url:
                continue
            products.append(TargetProduct(platform=platform, url=url, label=label or url))

    return pincodes, products
