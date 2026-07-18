"""Excel input/output for the keyword-search scraper.

Input:  an ``inputs.xlsx`` workbook — one sheet of product search terms,
        one sheet of pincodes — meant to be hand-edited in Excel by a
        non-technical user. ``main.py`` picks it up automatically when it
        exists, overriding the ``queries``/``pincodes`` lists in config.yaml.
        (Exact-SKU tracking has its own workbook: see targets.py.)

Output: a styled export workbook of everything in the SQLite database
        (optionally filtered), with a "Latest snapshot" sheet on top and
        the full "All observations" time series behind it. Served by the
        dashboard's Download Excel button and the export_excel.py CLI.
"""

from __future__ import annotations

import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# ------------------------------------------------------------------ #
# Input workbook
# ------------------------------------------------------------------ #

_PRODUCT_HEADERS = ("product", "product_name", "query", "search_term", "search term", "name")
_PINCODE_HEADERS = ("pincode", "pin_code", "pin code", "pin")


def _cell_str(value) -> str:
    """Normalize a cell to a clean string (Excel loves turning 700048 into 700048.0)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _read_column(ws: Worksheet, wanted_headers: tuple[str, ...]) -> list[str]:
    """Read one column of values: by header name if row 1 has a recognized
    header, else column A treating every row as data. Blanks skipped,
    order-preserving dedupe."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [_cell_str(c).lower() for c in rows[0]]
    col_idx = None
    for h in wanted_headers:
        if h in header:
            col_idx = header.index(h)
            break
    data_rows = rows if col_idx is None else rows[1:]
    if col_idx is None:
        col_idx = 0

    values: list[str] = []
    for row in data_rows:
        if col_idx >= len(row):
            continue
        v = _cell_str(row[col_idx])
        if v and v not in values:
            values.append(v)
    return values


def load_search_inputs(path: str | Path) -> tuple[list[str], list[str]]:
    """Read (queries, pincodes) from the inputs workbook.

    Sheets are found by name ("Products"/"Pincodes", case-insensitive prefix
    match) and otherwise by position: first non-Instructions sheet = products,
    second = pincodes — so a bare workbook with default "Sheet1"/"Sheet2"
    names works too.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    prod_ws = pin_ws = None
    for name in wb.sheetnames:
        low = name.strip().lower()
        if low.startswith("product") and prod_ws is None:
            prod_ws = wb[name]
        elif low.startswith(("pincode", "pin code", "pin_code")) and pin_ws is None:
            pin_ws = wb[name]

    positional = [
        wb[n] for n in wb.sheetnames
        if not n.strip().lower().startswith("instruction")
    ]
    if prod_ws is None and positional:
        prod_ws = positional[0]
    if pin_ws is None and len(positional) > 1:
        pin_ws = positional[1]

    queries = _read_column(prod_ws, _PRODUCT_HEADERS) if prod_ws is not None else []
    pincodes = _read_column(pin_ws, _PINCODE_HEADERS) if pin_ws is not None else []
    return queries, pincodes


# ------------------------------------------------------------------ #
# Output workbook
# ------------------------------------------------------------------ #

_EXPORT_COLUMNS = [
    # (sqlite column, spreadsheet header, width)
    ("timestamp", "Timestamp (UTC)", 22),
    ("platform", "Platform", 18),
    ("product_name", "Product", 42),
    ("brand", "Brand", 16),
    ("quantity_unit", "Pack size", 14),
    ("price", "Price (₹)", 12),
    ("mrp", "MRP (₹)", 12),
    ("discount_pct", "Discount %", 12),
    ("in_stock", "In stock", 10),
    ("stock_estimate", "Stock estimate", 14),
    ("stock_granularity", "Stock info type", 14),
    ("raw_stock_label", "Stock label", 20),
    ("pincode", "Pincode", 10),
    ("search_query", "Search term", 22),
    ("product_id", "Product ID", 24),
]


def _write_sheet(ws: Worksheet, rows: list[sqlite3.Row]) -> None:
    ws.append([header for _, header, _ in _EXPORT_COLUMNS])
    for i, (_, _, width) in enumerate(_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in rows:
        out = []
        for col, _, _ in _EXPORT_COLUMNS:
            value = row[col]
            if col == "in_stock":
                value = "Yes" if value else "No"
            out.append(value)
        ws.append(out)

    money_cols = [i for i, (c, _, _) in enumerate(_EXPORT_COLUMNS, start=1)
                  if c in ("price", "mrp", "discount_pct")]
    for r in range(2, ws.max_row + 1):
        for c in money_cols:
            ws.cell(row=r, column=c).number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_EXPORT_COLUMNS))}{max(ws.max_row, 1)}"


def build_export_workbook(
    db_path: str | Path,
    *,
    pincode: Optional[str] = None,
    query: Optional[str] = None,
    days: Optional[int] = None,
) -> tuple[Workbook, int]:
    """Build the export workbook; returns (workbook, observation row count)."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        where, params = [], []
        if pincode:
            where.append("pincode = ?")
            params.append(pincode)
        if query:
            where.append("search_query = ?")
            params.append(query)
        if days:
            cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat(
                timespec="seconds")
            where.append("timestamp >= ?")
            params.append(cutoff)
        clause = (" WHERE " + " AND ".join(where)) if where else ""
        cols = ", ".join(c for c, _, _ in _EXPORT_COLUMNS)

        latest = conn.execute(
            f"""
            SELECT {cols} FROM (
                SELECT *, ROW_NUMBER() OVER (
                    PARTITION BY platform, product_id, pincode
                    ORDER BY timestamp DESC) AS rn
                FROM observations{clause}
            ) WHERE rn = 1
            ORDER BY platform, product_name
            """,
            params,
        ).fetchall()
        all_rows = conn.execute(
            f"SELECT {cols} FROM observations{clause} ORDER BY timestamp DESC",
            params,
        ).fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws_latest = wb.active
    ws_latest.title = "Latest snapshot"
    _write_sheet(ws_latest, latest)
    _write_sheet(wb.create_sheet("All observations"), all_rows)
    return wb, len(all_rows)


def export_observations(
    db_path: str | Path,
    out_path: str | Path,
    *,
    pincode: Optional[str] = None,
    query: Optional[str] = None,
    days: Optional[int] = None,
) -> int:
    """Write the export workbook to ``out_path``; returns observation count."""
    wb, count = build_export_workbook(db_path, pincode=pincode, query=query, days=days)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return count
